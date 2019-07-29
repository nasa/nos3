import pytz
import re
import os
import warnings
warnings.filterwarnings('ignore', category=UserWarning, append=True)
from openpyxl import load_workbook
from datetime import date
from dateutil import parser

###############################################################################
# Python module to make it easy to work with a ground station tracking
# schedule excel spreadsheet 
###############################################################################

class GroundStationTrackingSchedule:
    def __init__(self, file_name="schedule.xlsm"):
        self.__file_name = file_name
        self.__worksheet = load_workbook(file_name).worksheets[0]
    
    # Alternative based on data in the spreadsheet
    def get_week(self):
        try:
            d = parser.parse(self.__worksheet.cell(row=6, column=10).value).date() # H6
        except:
            d = date(1, 1, 1)
        return d

    # Alternative based on filename
    #def get_week(self):
    #    #print(self.__file_name)
    #    result = re.search("(\d+)_(\d+)_(\d+)", self.__file_name)
    #    if (result is not None):
    #        #print("%s %s %s" % (result.group(3), result.group(1), result.group(2)))
    #        return date(int(result.group(3)), int(result.group(1)), int(result.group(2)))
    #    else:
    #        return date(1, 1, 1)
    
    # Alternative based on data in the spreadsheet
    #def get_revision(self):
    #    return self.__worksheet.cell(row=6, column=2).value # B6
    
    # Alternative based on filename
    #def get_revision(self):
    #    if (re.search("DRAFT", self.__file_name)):
    #        return "Draft"
    #    else:
    #        result = re.search("Rev_(\d+)", self.__file_name)
    #        if (result is not None):
    #            return "Final Rev %s" % result.group(1)
    #        else:
    #            return ""

    # Alternative based on last modified time 
    def get_revision(self):
        rev = os.path.getmtime(self.__file_name)
        #print(rev)
        return rev

    __revs = ["", "Draft", "Final Rev 0", "Final Rev 1", "Final Rev 2", "Final Rev 3", "Final Rev 4", \
        "Final Rev 5", "Final Rev 6", "Final Rev 7", "Final Rev 8", "Final Rev 9"]
    @classmethod
    # Alternative based on either the filename or data in the spreadsheet
    #def compare_revisions(self, rev1, rev2):
    #    r1 = self.__revs.index(rev1)
    #    #print("%s %s" % (rev1, r1))
    #    r2 = self.__revs.index(rev2)
    #    #print("%s %s" % (rev2, r2))
    #    if (r1 < r2):
    #        return -1
    #    elif (r1 == r2):
    #        return 0
    #    else:
    #        return 1

    # Alternative based on last modified time
    def compare_revisions(self, rev1, rev2):
        #print("%s %s" % (rev1, rev2))
        if (rev1 < rev2):
            return -1
        elif (rev1 == rev2):
            return 0
        else:
            return 1

    def get_satellite_contacts(self, sat_name):
        contacts = []
        for i in range(self.__worksheet.max_row):
            sat = self.__worksheet.cell(row=i+1, column=2)         # Satellite names found in column B
            support = self.__worksheet.cell(row=i+1, column=9)     # Y in column I means this satellite inview row is supported
            s = self.__worksheet.cell(row=i+1, column=3).value     # Column C - GMT start time of inview
            e = self.__worksheet.cell(row=i+1, column=4).value     # Column D - GMT end time of inview
            maxel = self.__worksheet.cell(row=i+1, column=5).value # Column E - maximum elevation of inview
            if ((sat.value == sat_name) and (support.value == 'Y')):
                start = pytz.UTC.localize(s)
                end   = pytz.UTC.localize(e)
                contacts.append((start, end, maxel))
        
        return contacts

